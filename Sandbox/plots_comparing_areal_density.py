# Script to plot AD2, AU32 and AU20 to show high and low RBS content.

import matplotlib.pyplot as plt
import openpyxl as xl
import numpy as np


def cell_to_value(cells_from_sheet):
    values_from_cells = np.zeros(len(cells_from_sheet))
    for index in range(0, len(cells_from_sheet)):
        cell = cells_from_sheet[index][0]
        rminrsep_or_areal = cell.value
        values_from_cells[index] = rminrsep_or_areal
    return values_from_cells

# Load workbooks.
pA2_wb  = xl.load_workbook("A2.xlsx", data_only=True)
pA32_wb = xl.load_workbook("A32.xlsx", data_only=True)
pA20_wb = xl.load_workbook("A20.xlsx", data_only=True)

# Get sheets.
pA2_sheet  = pA2_wb.get_sheet_by_name("A2")
pA32_sheet = pA32_wb.get_sheet_by_name("A32")
pA20_sheet = pA20_wb.get_sheet_by_name("A20")

# Get R-Rsep and areal cells. Returns tuple of tuples each with one cell in it.
pA2U_rminrsep_cells  = pA2_sheet["H2:H20"]
pA32D_rminrsep_cells = pA32_sheet["P2:P21"]
pA20D_rminrsep_cells = pA20_sheet["P2:P21"]
pA2U_areal_cells  = pA2_sheet["I2:I20"]
pA32D_areal_cells = pA32_sheet["J2:J21"]
pA20D_areal_cells = pA20_sheet["J2:J21"]
pA2U_arealerr_cells  = pA2_sheet["B2:B20"]
pA32D_arealerr_cells = pA32_sheet["O2:O21"]
pA20D_arealerr_cells = pA20_sheet["O2:O21"]

# New lists to hold values (instead of Cell objects).
pA2U_rminrsep  = cell_to_value(pA2U_rminrsep_cells)
pA32D_rminrsep = cell_to_value(pA32D_rminrsep_cells)
pA20D_rminrsep = cell_to_value(pA20D_rminrsep_cells)
pA2U_areal     = cell_to_value(pA2U_areal_cells)
pA32D_areal    = cell_to_value(pA32D_areal_cells)
pA20D_areal    = cell_to_value(pA20D_areal_cells)
pA2U_arealerr  = cell_to_value(pA2U_arealerr_cells)
pA32D_arealerr = cell_to_value(pA32D_arealerr_cells)
pA20D_arealerr = cell_to_value(pA20D_arealerr_cells)

# Create plot with each areal profile on it.
plt.errorbar(x=pA2U_rminrsep[:-1], y=pA2U_areal[:-1], yerr=pA2U_arealerr[:-1], capsize=3, fmt="g", label="AU2")
plt.errorbar(x=pA32D_rminrsep[:-1], y=pA32D_areal[:-1], yerr=pA32D_arealerr[:-1], capsize=3, fmt="y", label="AD32")
plt.errorbar(x=pA20D_rminrsep[:-1], y=pA20D_areal[:-1], yerr=pA20D_arealerr[:-1], capsize=3, fmt="r", label="AD20")
plt.xlabel(r"$\mathrm{R-R_{sep}\ (cm)}$")
plt.ylabel(r"$\mathrm{W\ Areal\ Density\ (10^{19}\ atoms/m^{2})}$")
plt.title("Comparison of Areal Densities")
plt.legend()
plt.show()

# BONUS PLOT. AU2 and AD2 on the same one. Just need AD2 data now.
pA2D_rminrsep_cells = pA2_sheet["F2:F20"]
pA2D_areal_cells    = pA2_sheet["C2:C20"]
pA2D_arealerr_cells = pA2_sheet["J2:J20"]
pA2D_rminrsep = cell_to_value(pA2D_rminrsep_cells)
pA2D_areal    = cell_to_value(pA2D_areal_cells)
pA2D_arealerr = cell_to_value(pA2D_arealerr_cells)
plt.errorbar(x=pA2U_rminrsep[:-1], y=pA2U_areal[:-1], yerr=pA2U_arealerr[:-1], capsize=3, label="AU2")
plt.errorbar(x=pA2D_rminrsep[:-1], y=pA2D_areal[:-1], yerr=pA2D_arealerr[:-1], capsize=3, label="AD2")
plt.xlabel(r"$\mathrm{R-R_{sep}\ (cm)}$")
plt.ylabel(r"$\mathrm{W\ Areal\ Density\ (10^{19}\ atoms/m^{2})}$")
plt.title("Comparison of Areal Densities")
plt.legend()
plt.show()
